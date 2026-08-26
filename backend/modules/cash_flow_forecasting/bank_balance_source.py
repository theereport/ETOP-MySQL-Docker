from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BANK_BALANCE_FILE_PATH = Path(
    r"F:\Accounting\Shared\Banking\Consolidated Daily Bank Balances.xlsx"
)

# Column positions on Sheet1, confirmed against the live file header row:
# A Date | ... | K Net Available | L LOC w/h - Travelers | M Line of
# Credit Balance | N Line of Credit Available
_DATE_COL = 1
_NET_AVAILABLE_COL = 11
_LOC_WITHHOLDING_COL = 12
_LOC_BALANCE_COL = 13
_LOC_AVAILABLE_COL = 14


class BankBalanceFileUnavailable(RuntimeError):
    """Raised when the shared banking workbook cannot be read."""


@dataclass(frozen=True)
class BankBalanceRow:
    business_day: date
    net_available: float | None
    line_of_credit_withholding: float | None
    line_of_credit_balance: float | None
    line_of_credit_available: float | None


def _as_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_bank_balance_rows(
    path: Path = BANK_BALANCE_FILE_PATH,
) -> list[BankBalanceRow]:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except (FileNotFoundError, OSError) as exc:
        raise BankBalanceFileUnavailable(
            f"Could not read the bank balance workbook at {path}: {exc}"
        ) from exc

    try:
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        rows: list[BankBalanceRow] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_date = row[_DATE_COL - 1] if len(row) >= _DATE_COL else None
            if isinstance(raw_date, datetime):
                business_day = raw_date.date()
            elif isinstance(raw_date, date):
                business_day = raw_date
            else:
                continue
            rows.append(
                BankBalanceRow(
                    business_day=business_day,
                    net_available=_as_float(
                        row[_NET_AVAILABLE_COL - 1]
                        if len(row) >= _NET_AVAILABLE_COL
                        else None
                    ),
                    line_of_credit_withholding=_as_float(
                        row[_LOC_WITHHOLDING_COL - 1]
                        if len(row) >= _LOC_WITHHOLDING_COL
                        else None
                    ),
                    line_of_credit_balance=_as_float(
                        row[_LOC_BALANCE_COL - 1]
                        if len(row) >= _LOC_BALANCE_COL
                        else None
                    ),
                    line_of_credit_available=_as_float(
                        row[_LOC_AVAILABLE_COL - 1]
                        if len(row) >= _LOC_AVAILABLE_COL
                        else None
                    ),
                )
            )
        return rows
    finally:
        workbook.close()


def latest_row_on_or_before(
    as_of: date, rows: list[BankBalanceRow] | None = None
) -> BankBalanceRow | None:
    """The most recent business-day row on or before `as_of`.

    The bank file is updated through the prior business day, so calling
    this with today's date correctly returns yesterday's (or last
    Friday's) row rather than a same-day figure that doesn't exist yet.
    """

    candidates = rows if rows is not None else read_bank_balance_rows()
    eligible = [row for row in candidates if row.business_day <= as_of]
    if not eligible:
        return None
    return max(eligible, key=lambda row: row.business_day)


def row_for_week_end(
    week_end: date, rows: list[BankBalanceRow] | None = None
) -> BankBalanceRow | None:
    """The actual ending balance for a closed week: the last business-day
    row on or before that week's Sunday."""

    return latest_row_on_or_before(week_end, rows)
