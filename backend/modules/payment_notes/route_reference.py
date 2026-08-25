"""Strict, provenance-preserving route-to-store reference import."""

from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass
from typing import Any, Iterable


ROUTE_REFERENCE_PARSER_VERSION = "payment-notes-route-reference@1.0.0"
REQUIRED_HEADERS = ("STORE", "ROUTE")
MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_MAPPING_ROWS = 5_000


class RouteReferenceError(ValueError):
    """The supplied route reference cannot be used safely."""


def normalize_route(value: Any) -> str:
    return str(value or "").strip().upper()


def normalize_store(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"[0-9]+(?:\.0+)?", text):
        digits = text.split(".", 1)[0]
        return str(int(digits)) if int(digits) else "0"
    return text.upper()


@dataclass(frozen=True)
class RouteReferenceData:
    source_name: str
    source_sha256: str
    source_size: int
    parser_version: str
    input_row_count: int
    blank_row_count: int
    duplicate_mapping_count: int
    raw_mappings: tuple[tuple[int, str, str], ...]
    mappings: tuple[tuple[str, str], ...]
    by_store: dict[str, tuple[str, ...]]
    conflicts: dict[str, tuple[str, ...]]
    warnings: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_name": self.source_name,
            "source_sha256": self.source_sha256,
            "source_size": self.source_size,
            "parser_version": self.parser_version,
            "input_row_count": self.input_row_count,
            "blank_row_count": self.blank_row_count,
            "duplicate_mapping_count": self.duplicate_mapping_count,
            "raw_mappings": [
                {"row_number": row_number, "store": store, "route": route}
                for row_number, store, route in self.raw_mappings
            ],
            "mappings": [
                {"store": store, "route": route}
                for store, route in self.mappings
            ],
            "by_store": {
                store: list(routes) for store, routes in self.by_store.items()
            },
            "conflicts": {
                route: list(stores) for route, stores in self.conflicts.items()
            },
            "warnings": list(self.warnings),
        }


@dataclass(frozen=True)
class RouteResolution:
    store_number: str
    routes: tuple[str, ...]
    conflicting_routes: tuple[str, ...]
    status: str


def _validated_headers(headers: Iterable[Any]) -> dict[str, int]:
    values = [str(value or "").strip() for value in headers]
    # XLSX worksheets frequently carry formatting into unused trailing
    # columns. Ignore only entirely blank trailing cells; an interior blank
    # header remains a structural error because it can hide populated data.
    while values and not values[-1]:
        values.pop()
    normalized = [value.upper() for value in values]
    if any(not value for value in normalized):
        raise RouteReferenceError("Route reference contains a blank header.")
    duplicates = sorted({value for value in normalized if normalized.count(value) > 1})
    if duplicates:
        raise RouteReferenceError(
            "Route reference contains duplicate headers: " + ", ".join(duplicates)
        )
    missing = [value for value in REQUIRED_HEADERS if value not in normalized]
    if missing:
        raise RouteReferenceError(
            "Route reference is missing required headers: " + ", ".join(missing)
        )
    return {normalized[index]: index for index in range(len(values))}


def _csv_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RouteReferenceError("Route reference CSV must be UTF-8.") from exc
    reader = csv.DictReader(io.StringIO(text, newline=""))
    if reader.fieldnames is None:
        raise RouteReferenceError("Route reference CSV has no header row.")
    header_positions = _validated_headers(reader.fieldnames)
    store_header = reader.fieldnames[header_positions["STORE"]]
    route_header = reader.fieldnames[header_positions["ROUTE"]]
    return [
        {
            "STORE": row.get(store_header),
            "ROUTE": row.get(route_header),
        }
        for row in reader
    ]


def _xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency exists in ETOP Lockbox
        raise RouteReferenceError(
            "XLSX route import requires the existing openpyxl Lockbox dependency."
        ) from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise RouteReferenceError("The route reference workbook could not be opened.") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration as exc:
            raise RouteReferenceError("Route reference workbook is empty.") from exc
        positions = _validated_headers(headers)
        return [
            {
                "STORE": values[positions["STORE"]]
                if positions["STORE"] < len(values)
                else None,
                "ROUTE": values[positions["ROUTE"]]
                if positions["ROUTE"] < len(values)
                else None,
            }
            for values in iterator
        ]
    finally:
        workbook.close()


def parse_route_reference(content: bytes, source_name: str) -> RouteReferenceData:
    if not content:
        raise RouteReferenceError("Route reference file is empty.")
    if len(content) > MAX_FILE_BYTES:
        raise RouteReferenceError(
            f"Route reference exceeds the {MAX_FILE_BYTES}-byte limit."
        )
    lower_name = source_name.lower()
    if lower_name.endswith(".csv"):
        rows = _csv_rows(content)
    elif lower_name.endswith(".xlsx"):
        rows = _xlsx_rows(content)
    else:
        raise RouteReferenceError("Route reference must be a .csv or .xlsx file.")

    if len(rows) > MAX_MAPPING_ROWS:
        raise RouteReferenceError(
            f"Route reference exceeds the {MAX_MAPPING_ROWS}-row limit."
        )
    pairs: set[tuple[str, str]] = set()
    raw_mappings: list[tuple[int, str, str]] = []
    blank_rows = 0
    for row_number, row in enumerate(rows, start=2):
        store = normalize_store(row.get("STORE"))
        route = normalize_route(row.get("ROUTE"))
        raw_store = str(row.get("STORE") or "")
        raw_route = str(row.get("ROUTE") or "")
        if not store and not route:
            blank_rows += 1
            continue
        if not store or not route:
            raise RouteReferenceError(
                f"Route reference row {row_number} must contain both STORE and ROUTE."
            )
        raw_mappings.append((row_number, raw_store, raw_route))
        pairs.add((store, route))
    if not pairs:
        raise RouteReferenceError("Route reference has no usable mappings.")

    route_stores: dict[str, set[str]] = {}
    for store, route in pairs:
        route_stores.setdefault(route, set()).add(store)
    conflicts = {
        route: tuple(sorted(stores))
        for route, stores in sorted(route_stores.items())
        if len(stores) > 1
    }
    by_store: dict[str, list[str]] = {}
    for store, route in sorted(pairs):
        if route in conflicts:
            continue
        by_store.setdefault(store, []).append(route)
    warnings: list[str] = []
    duplicate_count = len(rows) - blank_rows - len(pairs)
    if duplicate_count:
        warnings.append(f"Collapsed {duplicate_count} duplicate mapping row(s).")
    if blank_rows:
        warnings.append(f"Ignored {blank_rows} completely blank row(s).")
    if conflicts:
        warnings.append(
            f"Excluded {len(conflicts)} route code(s) mapped to multiple stores."
        )
    return RouteReferenceData(
        source_name=source_name,
        source_sha256=hashlib.sha256(content).hexdigest(),
        source_size=len(content),
        parser_version=ROUTE_REFERENCE_PARSER_VERSION,
        input_row_count=len(rows),
        blank_row_count=blank_rows,
        duplicate_mapping_count=duplicate_count,
        raw_mappings=tuple(raw_mappings),
        mappings=tuple(sorted(pairs)),
        by_store={store: tuple(routes) for store, routes in sorted(by_store.items())},
        conflicts=conflicts,
        warnings=tuple(warnings),
    )


def resolve_store(reference: RouteReferenceData | dict[str, Any], store: Any) -> RouteResolution:
    canonical_store = normalize_store(store)
    if isinstance(reference, RouteReferenceData):
        by_store = reference.by_store
        conflicts = reference.conflicts
    else:
        by_store = {
            key: tuple(value) for key, value in reference.get("by_store", {}).items()
        }
        conflicts = {
            key: tuple(value) for key, value in reference.get("conflicts", {}).items()
        }
    conflict_routes = tuple(
        sorted(route for route, stores in conflicts.items() if canonical_store in stores)
    )
    routes = tuple(sorted(by_store.get(canonical_store, ())))
    if not routes and conflict_routes:
        status = "conflict_only"
    elif not routes:
        status = "unmapped"
    elif conflict_routes:
        status = "mapped_with_excluded_conflicts"
    else:
        status = "mapped"
    return RouteResolution(
        store_number=canonical_store,
        routes=routes,
        conflicting_routes=conflict_routes,
        status=status,
    )


__all__ = [
    "ROUTE_REFERENCE_PARSER_VERSION",
    "RouteReferenceData",
    "RouteReferenceError",
    "RouteResolution",
    "normalize_route",
    "normalize_store",
    "parse_route_reference",
    "resolve_store",
]

