from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Num",
    "Env Num",
    "Envelope",
    "Transaction",
    "Lockbox",
    "Date",
    "Time",
    "Batch",
    "Batch Item",
    "Check",
    "Check Amount",
    "ABA/RT",
    "Account Num",
    "Check Num",
    "Invoice Number",
    "Net Invoice Amount",
    "Invoice Page",
    "Check Image",
    "Envelope Image",
    "Invoice Image",
]


def _excel_date(
    value: str,
) -> datetime | str:
    for fmt in (
        "%Y/%m/%d",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(
                value,
                fmt,
            )
        except ValueError:
            pass

    return value


def _copy_template_style(
    source,
    target,
) -> None:
    if source.has_style:
        target._style = copy(
            source._style
        )
    if source.number_format:
        target.number_format = (
            source.number_format
        )
    if source.alignment:
        target.alignment = copy(
            source.alignment
        )


def export_pnc_workbook(
    result: dict[str, Any],
    output_path: str | Path,
    template_path: str | Path | None = None,
) -> Path:
    output_path = Path(output_path)

    if (
        template_path
        and Path(template_path).exists()
    ):
        workbook = load_workbook(
            template_path
        )
        sheet = workbook["detail"]
        sheet.delete_rows(
            4,
            max(
                sheet.max_row - 3,
                1,
            ),
        )
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "detail"

        sheet["A1"] = " "
        sheet["A2"] = result.get(
            "lockbox",
            "",
        )
        sheet["B2"] = (
            "K and M Tire Inc"
        )
        sheet["C2"] = "Web Page"
        sheet["D2"] = "w1"
        sheet["E2"] = (
            result.get(
                "transaction_date",
                "",
            )
            .replace("/", "")
            .replace("-", "")
        )
        sheet["F2"] = (
            result.get(
                "transaction_count",
                0,
            )
        )
        sheet["G2"] = (
            result.get(
                "transaction_count",
                0,
            )
        )

        for column, header in enumerate(
            HEADERS,
            start=1,
        ):
            cell = sheet.cell(
                row=3,
                column=column,
                value=header,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor="1F4E78",
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    rows: list[list[Any]] = []
    sequence = 1
    check_sequence = 1

    for transaction in result.get(
        "transactions",
        [],
    ):
        allocations = transaction.get(
            "allocations",
            [],
        )

        if not allocations:
            allocations = [
                {
                    "invoice_number": "",
                    "net_invoice_amount":
                        0.0,
                    "invoice_page": "",
                }
            ]

        for allocation in allocations:
            rows.append(
                [
                    sequence,
                    transaction.get(
                        "envelope_number",
                    ),
                    transaction.get(
                        "transaction_id",
                    ),
                    transaction.get(
                        "transaction_id",
                    ),
                    transaction.get(
                        "lockbox",
                    ),
                    _excel_date(
                        transaction.get(
                            "date",
                            "",
                        )
                    ),
                    0.75,
                    transaction.get(
                        "batch",
                    ),
                    transaction.get(
                        "batch_item",
                    ),
                    check_sequence,
                    transaction.get(
                        "check_amount",
                        0.0,
                    ),
                    transaction.get(
                        "aba_routing",
                    ),
                    transaction.get(
                        "account_number",
                    ),
                    transaction.get(
                        "check_number",
                    ),
                    allocation.get(
                        "invoice_number",
                    ),
                    allocation.get(
                        "net_invoice_amount",
                        0.0,
                    ),
                    allocation.get(
                        "invoice_page",
                        "",
                    ),
                    transaction.get(
                        "check_page",
                    ),
                    "",
                    (
                        transaction.get(
                            "remittance_pages",
                            [""],
                        )[0]
                        if transaction.get(
                            "remittance_pages"
                        )
                        else ""
                    ),
                ]
            )
            sequence += 1

        check_sequence += 1

    start_row = 4

    for row_offset, row in enumerate(
        rows,
        start=start_row,
    ):
        for column, value in enumerate(
            row,
            start=1,
        ):
            target = sheet.cell(
                row=row_offset,
                column=column,
                value=value,
            )

            if (
                template_path
                and Path(
                    template_path
                ).exists()
                and sheet.max_row >= 4
            ):
                _copy_template_style(
                    sheet.cell(
                        row=4,
                        column=column,
                    ),
                    target,
                )

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = (
        f"A3:T{max(sheet.max_row, 3)}"
    )

    date_column = 6
    money_columns = [11, 16]

    for row in range(
        start_row,
        sheet.max_row + 1,
    ):
        sheet.cell(
            row=row,
            column=date_column,
        ).number_format = "mm/dd/yyyy"

        for column in money_columns:
            sheet.cell(
                row=row,
                column=column,
            ).number_format = (
                '$#,##0.00'
            )

    widths = {
        1: 8, 2: 10, 3: 15, 4: 15,
        5: 15, 6: 12, 7: 10, 8: 9,
        9: 11, 10: 9, 11: 14, 12: 13,
        13: 15, 14: 13, 15: 17,
        16: 18, 17: 14, 18: 13,
        19: 15, 20: 14,
    }

    for column, width in widths.items():
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = width

    workbook.save(output_path)
    return output_path
