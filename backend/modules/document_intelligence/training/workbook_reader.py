from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = {
    "Transaction",
    "Check Amount",
    "Invoice Number",
    "Net Invoice Amount",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _money(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    return round(float(cleaned or 0), 2)


def _date_text(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return _text(value)


def read_pnc_ground_truth(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook["detail"] if "detail" in workbook.sheetnames else workbook.active

    header_row = None
    headers: dict[str, int] = {}
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [_text(value) for value in row]
        candidate = {value: index for index, value in enumerate(values) if value}
        if REQUIRED_HEADERS.issubset(candidate):
            header_row = row_number
            headers = candidate
            break

    if header_row is None:
        raise ValueError(
            "The workbook does not contain the expected PNC detail headers: "
            + ", ".join(sorted(REQUIRED_HEADERS))
        )

    transactions: dict[str, dict[str, Any]] = {}
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        transaction_id = _text(row[headers["Transaction"]])
        if not transaction_id:
            continue
        transaction = transactions.setdefault(
            transaction_id,
            {
                "transaction_id": transaction_id,
                "check_amount": _money(row[headers["Check Amount"]]),
                "check_number": _text(row[headers.get("Check Num", -1)]) if "Check Num" in headers else "",
                "date": _date_text(row[headers.get("Date", -1)]) if "Date" in headers else "",
                "allocations": [],
            },
        )
        transaction["allocations"].append(
            {
                "invoice_number": _text(row[headers["Invoice Number"]]),
                "net_invoice_amount": _money(row[headers["Net Invoice Amount"]]),
            }
        )

    return {
        "transaction_count": len(transactions),
        "allocation_count": sum(len(item["allocations"]) for item in transactions.values()),
        "transactions": list(transactions.values()),
    }
