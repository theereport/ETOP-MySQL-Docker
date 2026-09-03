from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


QUEUE_HEADERS = (
    "Status",
    "Transaction",
    "Batch / Item",
    "Customer",
    "Check",
    "Check Amount",
    "Invoices",
    "Allocated",
    "Difference",
)

STATUS_LABELS = {
    "approved": "Approved",
    "held": "Held",
    "carryover": "Carryover",
    "corrected": "Corrected",
    "balanced": "Prepared",
    "no_remittance": "No Remittance",
    "review_required": "Review",
}

CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'


def _safe_file_part(value: Any, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "").strip())
    return cleaned.strip("_")[:80] or fallback


def export_review_queue_workbook(
    review: dict[str, Any],
    transaction_ids: list[str],
    queue_label: str,
    reason_code: str,
    output_dir: str | Path,
) -> Path:
    """Export a canonical, read-only projection of the visible review queue."""

    requested = [str(value or "").strip() for value in transaction_ids]
    if not requested or any(not value for value in requested):
        raise ValueError("At least one valid transaction is required for export.")
    if len(requested) != len(set(requested)):
        raise ValueError("The export transaction list contains duplicates.")

    transactions = {
        str(item.get("transaction_id") or ""): item
        for item in review.get("transactions", [])
    }
    missing = [value for value in requested if value not in transactions]
    if missing:
        raise ValueError(
            "The review queue changed before export; refresh and try again."
        )
    selected = [transactions[value] for value in requested]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Queue"
    sheet.sheet_view.showGridLines = False

    dark_fill = PatternFill("solid", fgColor="13243A")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    accent_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    label_font = Font(color="1F4E78", bold=True)
    thin_blue = Side(style="thin", color="B4C7DC")
    border = Border(bottom=thin_blue)

    sheet.merge_cells("A1:I1")
    title_cell = sheet["A1"]
    title_cell.value = "Lockbox Transaction Review Queue"
    title_cell.fill = dark_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=15)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26

    metadata = (
        ("Source File", review.get("source_file_name", "")),
        ("Lockbox", review.get("lockbox", "")),
        ("Processing Date", review.get("transaction_date", "")),
        ("Queue Scope", queue_label or "All transactions"),
        ("Reason Code", reason_code or "—"),
        ("Displayed Check Total", float(review.get("total_check_amount") or 0)),
        ("Exported UTC", datetime.now(timezone.utc).isoformat()),
        ("Exported Rows", len(selected)),
    )
    metadata_positions = (
        (2, 1), (2, 4), (2, 7), (3, 1),
        (3, 4), (3, 7), (4, 1), (4, 7),
    )
    for (label, value), (row, column) in zip(metadata, metadata_positions):
        label_cell = sheet.cell(row=row, column=column, value=label)
        label_cell.fill = accent_fill
        label_cell.font = label_font
        value_cell = sheet.cell(row=row, column=column + 1, value=value)
        if label == "Displayed Check Total":
            value_cell.number_format = CURRENCY_FORMAT

    header_row = 6
    for column, label in enumerate(QUEUE_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=label)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[header_row].height = 24

    for row_number, transaction in enumerate(selected, start=header_row + 1):
        status = str(transaction.get("status") or "review_required")
        batch = transaction.get("batch")
        item = transaction.get("batch_item")
        batch_item = (
            f"{batch if batch is not None else ''} / "
            f"{item if item is not None else ''}"
        )
        values = (
            STATUS_LABELS.get(status, status.replace("_", " ").title()),
            str(transaction.get("transaction_id") or ""),
            batch_item,
            str(transaction.get("customer_name") or ""),
            str(transaction.get("check_number") or ""),
            float(transaction.get("check_amount") or 0),
            len(transaction.get("allocations") or []),
            float(transaction.get("allocation_total") or 0),
            float(transaction.get("difference") or 0),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        for column in (6, 8, 9):
            sheet.cell(row=row_number, column=column).number_format = (
                CURRENCY_FORMAT
            )
        check_cell = sheet.cell(row=row_number, column=5)
        check_cell.number_format = "@"
        check_cell.quotePrefix = True
        sheet.cell(row=row_number, column=7).number_format = "0"

    final_row = header_row + len(selected)
    sheet.auto_filter.ref = f"A{header_row}:I{final_row}"
    sheet.freeze_panes = f"A{header_row + 1}"

    widths = (16, 20, 16, 38, 18, 18, 22, 18, 18)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    job_part = _safe_file_part(review.get("job_id"), "lockbox")
    scope_part = _safe_file_part(queue_label, "queue")
    output = output_root / f"{job_part}_Review_Queue_{scope_part}.xlsx"
    workbook.save(output)
    return output
