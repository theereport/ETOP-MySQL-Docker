from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import date, datetime, time as datetime_time
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import HTTPException

from sql_workspace import execute_mysql_query
from core.sql_validator import normalize_and_validate_sql

from data.mysql import get_engine, reports_table
from sqlalchemy import select

from .schemas import (
    AutomationDefinition,
    RunAutomationResponse,
)
from .validation import resolve_script_path


BACKEND_DIR = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_DIR = BACKEND_DIR / "data" / "automation_outputs"
AUTOMATION_ROW_LIMIT = int(
    os.getenv("SQL_AUTOMATION_ROW_LIMIT", "100000")
)


class AutomationExecutionError(RuntimeError):
    """Raised when an automation cannot be completed."""


def _render_tokens(
    value: str,
    automation_name: str,
    run_time: datetime,
) -> str:
    return (
        value
        .replace("{automation_name}", automation_name)
        .replace("{yyyy-MM-dd}", run_time.strftime("%Y-%m-%d"))
        .replace("{yyyyMMdd}", run_time.strftime("%Y%m%d"))
        .replace("{HHmm}", run_time.strftime("%H%M"))
        .replace("{run_date}", run_time.strftime("%Y-%m-%d"))
    )


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", value).strip()
    return cleaned or "automation-output"


def _get_output_path(
    automation: AutomationDefinition,
    run_time: datetime,
) -> Path:
    configured_folder = automation.delivery.output_folder.strip()

    output_folder = (
        Path(
            os.path.expandvars(
                os.path.expanduser(configured_folder)
            )
        )
        if configured_folder
        else DEFAULT_OUTPUT_DIR
    )

    if not output_folder.is_absolute():
        output_folder = BACKEND_DIR / output_folder

    output_folder.mkdir(parents=True, exist_ok=True)

    rendered_name = _render_tokens(
        automation.file_name_template,
        automation.name,
        run_time,
    )
    rendered_name = _safe_filename(rendered_name)

    extension = automation.output_format.strip().lower()
    if extension not in {"csv", "xlsx", "pdf"}:
        raise AutomationExecutionError(
            "Output format must be CSV, XLSX, or PDF."
        )

    current_suffix = Path(rendered_name).suffix.lower()
    desired_suffix = f".{extension}"

    if current_suffix != desired_suffix:
        rendered_name = f"{Path(rendered_name).stem}{desired_suffix}"

    return (output_folder / rendered_name).resolve()


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, datetime_time)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return str(value)


def _normalize_rows(
    rows: list[Any],
    columns: list[str],
) -> list[list[Any]]:
    normalized: list[list[Any]] = []

    for row in rows:
        if hasattr(row, "keys"):
            normalized.append(
                [_serialize_cell(row[column]) for column in columns]
            )
        else:
            normalized.append(
                [_serialize_cell(value) for value in row]
            )

    return normalized


def _write_csv(
    output_path: Path,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    with output_path.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as output_file:
        writer = csv.writer(output_file)
        writer.writerow(columns)
        writer.writerows(rows)


def _write_xlsx(
    output_path: Path,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise AutomationExecutionError(
            "Excel export requires openpyxl. Install it with: "
            "python -m pip install openpyxl"
        ) from exc

    workbook = Workbook(write_only=False)
    worksheet = workbook.active
    worksheet.title = "Results"
    worksheet.append(columns)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    sample_rows = rows[:500]
    for column_index, column_name in enumerate(columns, start=1):
        maximum_length = len(str(column_name))
        for row in sample_rows:
            if column_index <= len(row):
                maximum_length = max(
                    maximum_length,
                    len(str(row[column_index - 1])),
                )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(maximum_length + 2, 50)

    workbook.save(output_path)


def _write_pdf(
    output_path: Path,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise AutomationExecutionError(
            "PDF export requires reportlab. Install it with: "
            "python -m pip install reportlab"
        ) from exc

    maximum_pdf_rows = 5000
    if len(rows) > maximum_pdf_rows:
        raise AutomationExecutionError(
            "PDF export is limited to 5,000 rows. "
            "Use CSV or XLSX for larger results."
        )

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]
    body_style.fontSize = 7
    body_style.leading = 8

    table_data = [
        [Paragraph(str(column), body_style) for column in columns]
    ]
    for row in rows:
        table_data.append(
            [Paragraph(str(value), body_style) for value in row]
        )

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(letter),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    document.build([table])


def _write_query_output(
    automation: AutomationDefinition,
    result: dict[str, Any],
    run_time: datetime,
) -> Path:
    output_path = _get_output_path(automation, run_time)
    columns = list(result.get("columns", []))
    raw_rows = list(result.get("rows", []))
    normalized_rows = _normalize_rows(raw_rows, columns)

    output_format = automation.output_format.lower()
    if output_format == "csv":
        _write_csv(output_path, columns, normalized_rows)
    elif output_format == "xlsx":
        _write_xlsx(output_path, columns, normalized_rows)
    elif output_format == "pdf":
        _write_pdf(output_path, columns, normalized_rows)
    else:
        raise AutomationExecutionError(
            f"Unsupported output format: {output_format}"
        )

    return output_path


def _load_saved_report_sql(report_id: str) -> tuple[str, str]:
    if not report_id.strip():
        raise AutomationExecutionError(
            "Select a saved report before running this automation."
        )

    with get_engine().connect() as connection:
        row = connection.execute(
            select(reports_table.c.name, reports_table.c.sql_text).where(
                reports_table.c.id == report_id
            )
        ).mappings().first()

    if row is None:
        raise AutomationExecutionError(
            "The selected saved report could not be found."
        )

    report_name = str(row["name"])
    sql_text = str(row["sql_text"])

    if not sql_text.strip():
        raise AutomationExecutionError(
            f'The saved report "{report_name}" does not contain SQL.'
        )

    return report_name, sql_text


def _execute_sql_source(
    automation: AutomationDefinition,
    run_time: datetime,
) -> tuple[Path, int, str]:
    if automation.source_type == "sql":
        sql_text = automation.sql.strip()
        source_description = "SQL automation"

        if not sql_text:
            raise AutomationExecutionError(
                "Enter SQL before running this automation."
            )
    elif automation.source_type == "report":
        report_name, sql_text = _load_saved_report_sql(
            automation.report_id
        )
        source_description = f'Saved report "{report_name}"'
    else:
        raise AutomationExecutionError("Invalid SQL automation source.")

    try:
        validated_sql = normalize_and_validate_sql(sql_text)
        result = execute_mysql_query(
            validated_sql=validated_sql,
            row_limit=AUTOMATION_ROW_LIMIT,
            # Scheduled automations have no interactive user session to
            # attribute the run to - None is the "shared/legacy" bucket
            # visible to every SQL Workspace user, the correct home for a
            # system-triggered run rather than any one person's history.
            created_by=None,
        )
    except HTTPException as exc:
        raise AutomationExecutionError(str(exc.detail)) from exc
    except Exception as exc:
        raise AutomationExecutionError(
            f"SQL execution failed: {exc}"
        ) from exc

    output_path = _write_query_output(
        automation,
        result,
        run_time,
    )
    row_count = int(result.get("row_count", 0))

    message = (
        f"{source_description} completed. "
        f"{row_count:,} rows exported to {output_path.name}."
    )

    if result.get("limit_applied"):
        message += (
            " The configured automation limit of "
            f"{AUTOMATION_ROW_LIMIT:,} rows was applied."
        )

    return output_path, row_count, message


def _run_script(
    automation: AutomationDefinition,
) -> subprocess.CompletedProcess[str]:
    script_path = resolve_script_path(automation.script_path)

    if automation.source_type == "powershell":
        executable = (
            shutil.which("pwsh.exe")
            or shutil.which("powershell.exe")
            or shutil.which("pwsh")
            or shutil.which("powershell")
        )

        if not executable:
            raise AutomationExecutionError(
                "PowerShell could not be found on this computer."
            )

        command = [
            executable,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
        ]
    elif automation.source_type == "python":
        command = [sys.executable, str(script_path)]
    else:
        raise AutomationExecutionError(
            "Invalid script automation source."
        )

    try:
        result = subprocess.run(
            command,
            cwd=str(script_path.parent),
            capture_output=True,
            text=True,
            timeout=3600,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise AutomationExecutionError(
            "The automation exceeded the one-hour execution limit."
        ) from exc
    except OSError as exc:
        raise AutomationExecutionError(
            f"The script could not be started: {exc}"
        ) from exc

    if result.returncode != 0:
        details = (
            result.stderr.strip()
            or result.stdout.strip()
            or f"Process returned exit code {result.returncode}."
        )
        raise AutomationExecutionError(
            f"Automation script failed: {details[-4000:]}"
        )

    return result


def _candidate_output_paths(stdout: str) -> list[Path]:
    candidates: list[Path] = []
    patterns = [
        r"(?im)^\s*OUTPUT_FILE\s*=\s*(.+?)\s*$",
        r"(?im)^\s*REPORT_PATH\s*=\s*(.+?)\s*$",
        r"(?im)^\s*Report created:\s*(.+?)\s*$",
        r"(?im)^\s*Output file:\s*(.+?)\s*$",
    ]

    for pattern in patterns:
        for match in re.findall(pattern, stdout):
            cleaned = match.strip().strip('"').strip("'")
            candidates.append(Path(cleaned))

    return candidates


def _find_script_output_file(
    automation: AutomationDefinition,
    stdout: str,
    run_time: datetime,
) -> Path | None:
    for candidate in _candidate_output_paths(stdout):
        expanded = Path(
            os.path.expandvars(
                os.path.expanduser(str(candidate))
            )
        )
        if expanded.exists() and expanded.is_file():
            return expanded.resolve()

    expected = _get_output_path(automation, run_time)
    if expected.exists() and expected.is_file():
        return expected

    return None


def _send_outlook_email(
    automation: AutomationDefinition,
    output_file: Path | None,
    run_time: datetime,
) -> None:
    delivery = automation.delivery

    if not delivery.recipients:
        raise AutomationExecutionError(
            "Email delivery is selected, but no recipient email "
            "address was entered."
        )

    executable = (
        shutil.which("powershell.exe")
        or shutil.which("pwsh.exe")
        or shutil.which("powershell")
        or shutil.which("pwsh")
    )

    if not executable:
        raise AutomationExecutionError(
            "PowerShell could not be found for Outlook email delivery."
        )

    subject = _render_tokens(
        delivery.subject or automation.name,
        automation.name,
        run_time,
    )
    body = _render_tokens(
        delivery.message or "The automation completed successfully.",
        automation.name,
        run_time,
    )

    attachment = (
        str(output_file)
        if output_file is not None and delivery.attach_output
        else ""
    )

    script = r'''
$ErrorActionPreference = "Stop"

$outlook = New-Object -ComObject Outlook.Application
$mail = $outlook.CreateItem(0)

$mail.To = $env:ETOP_MAIL_TO
$mail.CC = $env:ETOP_MAIL_CC
$mail.Subject = $env:ETOP_MAIL_SUBJECT
$mail.Body = $env:ETOP_MAIL_BODY

if (
    $env:ETOP_MAIL_ATTACHMENT -and
    (Test-Path -LiteralPath $env:ETOP_MAIL_ATTACHMENT)
) {
    [void]$mail.Attachments.Add(
        $env:ETOP_MAIL_ATTACHMENT
    )
}

$mail.Send()
'''

    environment = os.environ.copy()
    environment.update(
        {
            "ETOP_MAIL_TO": ";".join(delivery.recipients),
            "ETOP_MAIL_CC": ";".join(delivery.cc_recipients),
            "ETOP_MAIL_SUBJECT": subject,
            "ETOP_MAIL_BODY": body,
            "ETOP_MAIL_ATTACHMENT": attachment,
        }
    )

    with tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".ps1",
        encoding="utf-8",
        delete=False,
    ) as temp_file:
        temp_file.write(script)
        temp_path = Path(temp_file.name)

    try:
        result = subprocess.run(
            [
                executable,
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(temp_path),
            ],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
            env=environment,
        )
    finally:
        temp_path.unlink(missing_ok=True)

    if result.returncode != 0:
        details = (
            result.stderr.strip()
            or result.stdout.strip()
            or "Outlook returned an unknown error."
        )
        raise AutomationExecutionError(
            "The automation ran, but Outlook email delivery failed: "
            f"{details[-4000:]}"
        )


def run_automation(
    automation: AutomationDefinition,
) -> RunAutomationResponse:
    started = time.perf_counter()
    run_time = datetime.now()

    output_file: Path | None = None
    row_count: int | None = None
    warning_messages: list[str] = []

    if automation.source_type in {"sql", "report"}:
        output_file, row_count, message = _execute_sql_source(
            automation,
            run_time,
        )
    elif automation.source_type in {"powershell", "python"}:
        result = _run_script(automation)
        output_file = _find_script_output_file(
            automation,
            result.stdout,
            run_time,
        )
        script_message = result.stdout.strip()
        message = (
            script_message[-1500:]
            if script_message
            else "Automation completed successfully."
        )
    else:
        raise AutomationExecutionError(
            "Unsupported automation source type: "
            f"{automation.source_type}"
        )

    if automation.delivery.method == "email":
        _send_outlook_email(
            automation,
            output_file,
            run_time,
        )

        if automation.delivery.attach_output and output_file is None:
            warning_messages.append(
                "The email was sent without an attachment because "
                "no output file was detected."
            )
    elif automation.delivery.method == "folder":
        if output_file is None:
            warning_messages.append(
                "Folder delivery was selected, but no output file "
                "was detected."
            )

    if warning_messages:
        message = f"{message} {' '.join(warning_messages)}".strip()

    duration_ms = round((time.perf_counter() - started) * 1000)

    return RunAutomationResponse(
        status="warning" if warning_messages else "success",
        duration_ms=duration_ms,
        row_count=row_count,
        output_file_name=output_file.name if output_file else "",
        output_file_path=str(output_file) if output_file else "",
        message=message,
    )
