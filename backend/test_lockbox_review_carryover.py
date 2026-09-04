from __future__ import annotations

import tempfile
import unittest
from copy import deepcopy
from pathlib import Path
from unittest.mock import patch

from sqlalchemy import create_engine

from data.mysql import _reset_engine_override, _set_engine_override
from modules.document_intelligence.lockbox_review import service as review_service
from modules.document_intelligence.lockbox_review.schemas import (
    SaveTransactionReviewRequest,
)


class FakePayerMappingRepository:
    def __init__(self, *_args, **_kwargs) -> None:
        pass

    def upsert(self, *_args, **_kwargs) -> None:
        pass


def _source(*transactions: dict) -> dict:
    return {"transactions": list(transactions)}


class CarryoverStatusTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.engine = create_engine(
            f"sqlite:///{Path(self._tmpdir.name) / 'review.db'}"
        )
        _set_engine_override(self.engine)
        self.addCleanup(_reset_engine_override)
        self.addCleanup(self.engine.dispose)
        self._patches = [
            patch.object(
                review_service,
                "PayerCustomerMappingRepository",
                FakePayerMappingRepository,
            ),
            patch.object(review_service, "_governed_preparation_loader", None),
            patch.object(review_service, "_current_open_ar_loader", None),
        ]
        for patcher in self._patches:
            patcher.start()
            self.addCleanup(patcher.stop)

    def _transaction(self, transaction_id: str, **overrides) -> dict:
        base = {
            "transaction_id": transaction_id,
            "check_amount": 100.00,
            "aba_routing": "076401251",
            "account_number": "998877661234",
            "original_allocations": [],
            "allocations": [],
        }
        base.update(overrides)
        return base

    def _payload(self, **overrides) -> dict:
        base = SaveTransactionReviewRequest(
            allocations=[],
            expected_reviewed_at=None,
            status="carryover",
            customer_number="640194",
            customer_name="Gothenburg Tire",
        ).model_dump()
        base.update(overrides)
        return base

    def test_carryover_does_not_require_an_invoice_number(self) -> None:
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda _job_id: deepcopy(_source(self._transaction("T-1"))),
        ):
            result = review_service.save_transaction_review(
                "job-1", "T-1", self._payload(),
            )

        transaction = result["transactions"][0]
        self.assertEqual(transaction["status"], "carryover")

    def test_carryover_is_excluded_from_review_count(self) -> None:
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda _job_id: deepcopy(
                _source(self._transaction("T-1"), self._transaction("T-2"))
            ),
        ):
            result = review_service.save_transaction_review(
                "job-2", "T-1", self._payload(),
            )

        self.assertEqual(result["carryover_count"], 1)
        # T-2 is untouched (no allocations, no saved status) and still counts
        # as an open exception; T-1 (carryover) must not inflate that count.
        self.assertEqual(result["review_count"], 1)

    def test_reviewed_export_excludes_carryover_and_recomputes_totals(self) -> None:
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda _job_id: deepcopy(
                _source(
                    self._transaction("T-1", check_amount=100.00),
                    self._transaction("T-2", check_amount=50.00),
                )
            ),
        ):
            review_service.save_transaction_review(
                "job-3", "T-1", self._payload(),
            )
            review_service.save_transaction_review(
                "job-3",
                "T-2",
                self._payload(
                    status="corrected",
                    allocations=[
                        {"invoice_number": "12345678", "net_invoice_amount": 50.00}
                    ],
                ),
            )
            exported = review_service.build_reviewed_result("job-3")

        exported_ids = [item["transaction_id"] for item in exported["transactions"]]
        self.assertEqual(exported_ids, ["T-2"])
        self.assertEqual(exported["transaction_count"], 1)
        self.assertEqual(exported["total_check_amount"], 50.00)
        self.assertEqual(exported["total_allocation_amount"], 50.00)
        self.assertEqual(exported["total_difference"], 0.0)


class CrossJobCarryoverTest(unittest.TestCase):
    """list_carryover_transactions/create_carryover_export - the Carryover
    Dashboard's cross-job aggregation, distinct from the per-job save/
    export tests above."""

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.engine = create_engine(
            f"sqlite:///{Path(self._tmpdir.name) / 'review.db'}"
        )
        _set_engine_override(self.engine)
        self.addCleanup(_reset_engine_override)
        self.addCleanup(self.engine.dispose)
        self._patches = [
            patch.object(
                review_service,
                "PayerCustomerMappingRepository",
                FakePayerMappingRepository,
            ),
            patch.object(review_service, "_governed_preparation_loader", None),
            patch.object(review_service, "_current_open_ar_loader", None),
        ]
        for patcher in self._patches:
            patcher.start()
            self.addCleanup(patcher.stop)

    def _transaction(self, transaction_id: str, **overrides) -> dict:
        base = {
            "transaction_id": transaction_id,
            "check_amount": 100.00,
            "aba_routing": "076401251",
            "account_number": "998877661234",
            "original_allocations": [],
            "allocations": [],
        }
        base.update(overrides)
        return base

    def _payload(self, **overrides) -> dict:
        base = SaveTransactionReviewRequest(
            allocations=[],
            expected_reviewed_at=None,
            status="carryover",
            customer_number="640194",
            customer_name="Gothenburg Tire",
        ).model_dump()
        base.update(overrides)
        return base

    def test_list_carryover_transactions_spans_multiple_jobs(self) -> None:
        sources = {
            "job-a": _source(self._transaction("A-1")),
            "job-b": _source(self._transaction("B-1"), self._transaction("B-2")),
        }
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda job_id: deepcopy(sources[job_id]),
        ):
            review_service.save_transaction_review(
                "job-a", "A-1", self._payload(),
            )
            review_service.save_transaction_review(
                "job-b", "B-1", self._payload(),
            )
            review_service.save_transaction_review(
                "job-b",
                "B-2",
                self._payload(
                    status="corrected",
                    allocations=[
                        {"invoice_number": "12345678", "net_invoice_amount": 100.00}
                    ],
                ),
            )
            listed = review_service.list_carryover_transactions()

        listed_pairs = sorted(
            (item["job_id"], item["transaction_id"]) for item in listed
        )
        self.assertEqual(listed_pairs, [("job-a", "A-1"), ("job-b", "B-1")])

    def test_create_carryover_export_only_includes_approved_carryover_origin(
        self,
    ) -> None:
        sources = {
            "job-a": _source(self._transaction("A-1", check_amount=75.00)),
            "job-b": _source(self._transaction("B-1", check_amount=25.00)),
        }
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda job_id: deepcopy(sources[job_id]),
        ):
            # A-1: carried over, then approved - belongs in the export.
            carried_over = review_service.save_transaction_review(
                "job-a", "A-1", self._payload(),
            )
            carried_over_at = next(
                item["reviewed_at"]
                for item in carried_over["transactions"]
                if item["transaction_id"] == "A-1"
            )
            review_service.save_transaction_review(
                "job-a",
                "A-1",
                self._payload(
                    status="approved",
                    expected_reviewed_at=carried_over_at,
                    allocations=[
                        {"invoice_number": "12345678", "net_invoice_amount": 75.00}
                    ],
                ),
            )
            # B-1: approved directly, never carried over - must be excluded.
            review_service.save_transaction_review(
                "job-b",
                "B-1",
                self._payload(
                    status="approved",
                    allocations=[
                        {"invoice_number": "87654321", "net_invoice_amount": 25.00}
                    ],
                ),
            )
            output = review_service.create_carryover_export()

        self.assertTrue(output.exists())
        from openpyxl import load_workbook
        workbook = load_workbook(output)
        sheet = workbook["detail"]
        # Column D ("Transaction", HEADERS[3]) holds transaction_id.
        transaction_ids = [
            row[3].value for row in sheet.iter_rows(min_row=4)
            if row[3].value
        ]
        self.assertEqual(transaction_ids, ["A-1"])

    def test_create_carryover_export_can_be_scoped_to_one_customer(
        self,
    ) -> None:
        sources = {
            "job-a": _source(self._transaction("A-1", check_amount=75.00)),
            "job-b": _source(self._transaction("B-1", check_amount=25.00)),
        }
        with patch.object(
            review_service,
            "get_lockbox_result",
            lambda job_id: deepcopy(sources[job_id]),
        ):
            # A-1 and B-1 are both approved-after-carryover, but for two
            # different customers.
            carried_over_a = review_service.save_transaction_review(
                "job-a", "A-1", self._payload(customer_number="640194"),
            )
            review_service.save_transaction_review(
                "job-a",
                "A-1",
                self._payload(
                    customer_number="640194",
                    status="approved",
                    expected_reviewed_at=next(
                        item["reviewed_at"]
                        for item in carried_over_a["transactions"]
                        if item["transaction_id"] == "A-1"
                    ),
                    allocations=[
                        {"invoice_number": "12345678", "net_invoice_amount": 75.00}
                    ],
                ),
            )
            carried_over_b = review_service.save_transaction_review(
                "job-b", "B-1", self._payload(customer_number="999999"),
            )
            review_service.save_transaction_review(
                "job-b",
                "B-1",
                self._payload(
                    customer_number="999999",
                    status="approved",
                    expected_reviewed_at=next(
                        item["reviewed_at"]
                        for item in carried_over_b["transactions"]
                        if item["transaction_id"] == "B-1"
                    ),
                    allocations=[
                        {"invoice_number": "87654321", "net_invoice_amount": 25.00}
                    ],
                ),
            )
            output = review_service.create_carryover_export("640194")

        from openpyxl import load_workbook
        workbook = load_workbook(output)
        sheet = workbook["detail"]
        transaction_ids = [
            row[3].value for row in sheet.iter_rows(min_row=4)
            if row[3].value
        ]
        self.assertEqual(transaction_ids, ["A-1"])
        self.assertIn("640194", output.name)


if __name__ == "__main__":
    unittest.main()
