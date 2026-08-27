import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from modules.document_intelligence.pnc_lockbox_export import export_pnc_workbook


class PncLockboxExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()

    def tearDown(self) -> None:
        self._tmpdir.cleanup()

    def _export(self, result):
        output_path = Path(self._tmpdir.name) / "export.xlsx"
        export_pnc_workbook(result, output_path)
        return load_workbook(output_path)["detail"]

    def test_customer_number_header_is_column_u(self) -> None:
        sheet = self._export({"transactions": []})

        self.assertEqual(sheet["U3"].value, "Customer Number")

    def test_customer_number_value_is_written_per_allocation_row(self) -> None:
        sheet = self._export(
            {
                "transactions": [
                    {
                        "transaction_id": "G-1",
                        "customer_number": "104014",
                        "allocations": [
                            {"invoice_number": "841589278", "net_invoice_amount": 100.0},
                            {"invoice_number": "841594500", "net_invoice_amount": 50.0},
                        ],
                    },
                ],
            }
        )

        self.assertEqual(sheet["U4"].value, "104014")
        self.assertEqual(sheet["U5"].value, "104014")

    def test_missing_customer_number_is_blank(self) -> None:
        sheet = self._export(
            {
                "transactions": [
                    {
                        "transaction_id": "G-1",
                        "allocations": [
                            {"invoice_number": "841589278", "net_invoice_amount": 100.0},
                        ],
                    },
                ],
            }
        )

        self.assertIn(sheet["U4"].value, (None, ""))

    def test_misc_gl_sheet_lists_only_entries_with_an_amount(self) -> None:
        output_path = Path(self._tmpdir.name) / "export.xlsx"
        export_pnc_workbook(
            {
                "transactions": [
                    {
                        "transaction_id": "G-1",
                        "check_number": "394800",
                        "customer_number": "104014",
                        "allocations": [
                            {"invoice_number": "841589278", "net_invoice_amount": 90.0},
                        ],
                        "misc_gl": {
                            "reason": "Service Charge ADJ",
                            "gl_code": "3880",
                            "location": "DAL",
                            "department": "AR",
                            "amount": 10.0,
                        },
                    },
                    {
                        "transaction_id": "G-2",
                        "check_number": "111111",
                        "customer_number": "200000",
                        "allocations": [
                            {"invoice_number": "841594500", "net_invoice_amount": 50.0},
                        ],
                    },
                ],
            },
            output_path,
        )
        workbook = load_workbook(output_path)

        self.assertIn("misc_gl", workbook.sheetnames)
        sheet = workbook["misc_gl"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Check #", "Customer #", "GL Code", "Location", "Department", "Amount"],
        )
        self.assertEqual(
            [sheet.cell(row=2, column=col).value for col in range(1, 7)],
            ["394800", "104014", "3880", "DAL", "AR", 10.0],
        )
        self.assertIsNone(sheet.cell(row=3, column=1).value)


if __name__ == "__main__":
    unittest.main()
