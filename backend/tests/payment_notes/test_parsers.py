import csv
from copy import copy
import io
from decimal import Decimal
import unittest
from openpyxl import Workbook

from modules.payment_notes.remote_capture import REQUIRED_HEADERS, parse_remote_capture, physical_items
from modules.payment_notes.route_reference import parse_route_reference, resolve_store


def capture(rows):
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(REQUIRED_HEADERS)
    writer.writerows(rows)
    return stream.getvalue().encode()


def row(item_type, amount, check="0007", channel="Warehouse", store="1", deposit="D-1"):
    return ["08/20/2026 10:00:00 AM", "", "", f"{store} - Example", "", "Synthetic",
            deposit, item_type, amount, "", "", check, channel]


class ParserTests(unittest.TestCase):
    def test_balancing_uses_virtual_credit_only_and_keeps_blank_check_in_totals(self):
        parsed = parse_remote_capture(capture([
            row("Business Check", "10.00", ""), row("Personal Check", "5.00", "0007"),
            row("Virtual Credit", "15.00", "")]), "synthetic.csv")
        self.assertEqual(len(physical_items(parsed, "run")), 2)
        self.assertEqual(parsed.deposits[0].location_key, "L01")
        self.assertEqual(parsed.deposits[0].status, "BALANCED")
        self.assertEqual(parsed.deposits[0].physical_total, Decimal("15.00"))
        self.assertIn("CHECK_NUMBER_MISSING", parsed.rows[0].warnings)

    def test_missing_optional_trailing_fields_are_padded_with_explicit_warnings(self):
        short = row("Business Check", "8.25")[:-2]
        parsed = parse_remote_capture(capture([short, row("Business Check", "1.75", "8"),
                                                row("Virtual Credit", "10.00", "")]), "synthetic.csv")
        self.assertEqual(len(parsed.quarantined_rows), 0)
        self.assertIn("SOURCE_TRAILING_OPTIONAL_COLUMNS_PADDED", parsed.rows[0].warnings)
        self.assertIn("CHECK_NUMBER_MISSING", parsed.rows[0].warnings)
        self.assertIn("CAPTURE_CHANNEL_BLANK", parsed.rows[0].warnings)
        deposit = parsed.deposits[0]
        self.assertEqual(deposit.physical_item_count, 2)
        self.assertEqual(deposit.physical_total, Decimal("10.00"))
        self.assertEqual(deposit.difference, Decimal("0.00"))
        self.assertEqual(deposit.status, "BALANCED")

    def test_missing_required_trailing_field_still_quarantines(self):
        missing_routing_number_and_after = row("Business Check", "8.25")[:-3]
        parsed = parse_remote_capture(
            capture([missing_routing_number_and_after]),
            "synthetic.csv",
        )
        self.assertEqual(len(parsed.rows), 0)
        self.assertEqual(len(parsed.quarantined_rows), 1)
        self.assertIn(
            "SOURCE_COLUMN_COUNT_MISMATCH",
            parsed.quarantined_rows[0].reason_codes,
        )

    def test_blank_capture_channel_warns_and_unknown_type_quarantines(self):
        parsed = parse_remote_capture(capture([row("Business Check", "2.00", channel=""),
                                                row("Virtual Credit", "2.00", ""),
                                                row("Coupon", "1.00")]), "synthetic.csv")
        self.assertIn("CAPTURE_CHANNEL_BLANK", parsed.rows[0].warnings)
        self.assertIn("ITEM_TYPE_UNSUPPORTED", parsed.quarantined_rows[0].reason_codes)

    def test_route_reference_conflict_is_explicit_and_excluded(self):
        parsed = parse_route_reference(b" STORE , route \n22,AA\n89,ZZ\n90,ZZ\n", "routes.csv")
        self.assertEqual(resolve_store(parsed, "22").routes, ("AA",))
        conflict = resolve_store(parsed, "89")
        self.assertEqual(conflict.status, "conflict_only")
        self.assertEqual(conflict.conflicting_routes, ("ZZ",))

    def test_xlsx_ignores_styled_trailing_blank_columns(self):
        workbook = Workbook(); sheet = workbook.active
        sheet.append([" STORE ", " route ", None, None])
        sheet.append([22, "AA", None, None])
        sheet.cell(row=1, column=4).fill = copy(sheet.cell(row=1, column=1).fill)
        stream = io.BytesIO(); workbook.save(stream); workbook.close()
        parsed = parse_route_reference(stream.getvalue(), "routes.xlsx")
        self.assertEqual(parsed.by_store["22"], ("AA",))
