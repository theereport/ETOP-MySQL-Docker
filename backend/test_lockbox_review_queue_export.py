from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from modules.document_intelligence.lockbox_review.queue_export import (
    export_review_queue_workbook,
)


class LockboxReviewQueueExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.review = {
            "job_id": "job-123",
            "source_file_name": "8.14.26P.pdf",
            "lockbox": "P-GH-640045",
            "transaction_date": "2026/08/14",
            "total_check_amount": 884540.64,
            "transactions": [
                {
                    "transaction_id": "G-2",
                    "status": "no_remittance",
                    "batch": 2,
                    "batch_item": 4,
                    "customer_name": "Second Customer",
                    "check_number": "000222",
                    "check_amount": 1200.25,
                    "allocations": [],
                    "allocation_total": 0,
                    "difference": 1200.25,
                },
                {
                    "transaction_id": "G-1",
                    "status": "review_required",
                    "batch": 1,
                    "batch_item": 3,
                    "customer_name": "First Customer",
                    "check_number": "000111",
                    "check_amount": 100,
                    "allocations": [{"invoice_number": "1"}],
                    "allocation_total": 96,
                    "difference": 4,
                },
            ],
        }

    def test_exports_requested_canonical_rows_in_requested_order(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            output = export_review_queue_workbook(
                self.review,
                ["G-1", "G-2"],
                "No reliable customer found",
                "no_reliable_customer",
                folder,
            )

            self.assertTrue(output.exists())
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Review Queue"]
            self.assertEqual(
                [sheet.cell(6, column).value for column in range(1, 10)],
                [
                    "Status", "Transaction", "Batch / Item", "Customer",
                    "Check", "Check Amount", "Invoices", "Allocated",
                    "Difference",
                ],
            )
            self.assertEqual(sheet["B3"].value, "No reliable customer found")
            self.assertEqual(sheet["E3"].value, "no_reliable_customer")
            self.assertEqual(sheet["H3"].value, 884540.64)
            self.assertEqual(sheet["B7"].value, "G-1")
            self.assertEqual(sheet["E7"].value, "000111")
            self.assertEqual(sheet["E7"].number_format, "@")
            self.assertTrue(sheet["E7"].quotePrefix)
            self.assertEqual(sheet["F7"].value, 100)
            self.assertEqual(sheet["G7"].value, 1)
            self.assertEqual(sheet["B8"].value, "G-2")
            self.assertEqual(sheet["A8"].value, "No Remittance")

    def test_rejects_stale_or_duplicate_transaction_projection(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            with self.assertRaisesRegex(ValueError, "changed before export"):
                export_review_queue_workbook(
                    self.review,
                    ["missing"],
                    "Needs Review",
                    "",
                    folder,
                )
            with self.assertRaisesRegex(ValueError, "duplicates"):
                export_review_queue_workbook(
                    self.review,
                    ["G-1", "G-1"],
                    "Needs Review",
                    "",
                    folder,
                )


if __name__ == "__main__":
    unittest.main()
